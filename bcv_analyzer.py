import contextlib
import csv
import itertools
import json
import os
import sys
import threading
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Annotated, Any, Callable
from urllib.parse import urlparse

import questionary
import typer
from rich.console import Console
from rich.panel import Panel


ALLOWED_TABLES = ("request", "ad", "slot", "candidate", "auction", "ack")
SRC_CATALOG = "mrm_log_flat"
SRC_SCHEMA = "default"
BCV_CATALOG = "etl"
BCV_SCHEMA = "public_test1"
EXCLUDED_DESCRIBE_COLUMNS = {"extra", "comment"}
UI_CONSOLE = Console()
_SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = _SCRIPT_DIR / "output"
FIELD_SIZE_DIR = _SCRIPT_DIR / "field_size"
ETL_FIELDS_PATH = _SCRIPT_DIR / "etl_fields.json"
APP_VERSION = "v0.1"
USAGE_COLUMNS = ("usage:ETL", "usage:Insights", "usage:Arena", "usage:LQS", "usage:Others")
USAGE_QUERY_LIMIT = 10
USAGE_QUERY_BATCH_SIZE = 100
USAGE_QUERY_MAX_RETRIES = 3


class RunMode(str, Enum):
    TRIAL = "Trial"
    FULL_RUN = "Full Run"


APP_BANNER = r"""
 ____   ____ __     __     _                _
| __ ) / ___|\ \   / /    / \   _ __   ___ | |_   _ _______ _ __
|  _ \| |     \ \ / /    / _ \ | '_ \ / _ \| | | | |_  / _ \ '__|
| |_) | |___   \ V /    / ___ \| | | | (_| | | |_| |/ /  __/ |
|____/ \____|   \_/    /_/   \_\_| |_|\__,_|_|\__, /___\___|_|
                                              |___/
"""


def build_http_headers(auth_token: str | None, auth_header: str | None) -> dict[str, str]:
    if not auth_token:
        return {}

    header_name = auth_header or "Authorization"
    if header_name.lower() == "authorization":
        return {header_name: f"Bearer {auth_token}"}
    return {header_name: auth_token}


def current_timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def log_info(message: str) -> None:
    print(f"[{current_timestamp()}] {message}")


def log_progress(message: str, done: bool = False) -> None:
    print(f"\r[{current_timestamp()}] {message}", end="\n" if done else "", flush=True)


@contextlib.contextmanager
def spinning_cursor(message: str = "Executing Presto query"):
    """Display a spinning cursor animation while the enclosed block runs."""
    stop_event = threading.Event()
    frames = itertools.cycle(["|", "/", "-", "\\"])

    def _spin() -> None:
        while not stop_event.is_set():
            print(f"\r[{current_timestamp()}] {message} {next(frames)}", end="", flush=True)
            stop_event.wait(timeout=1.0)
        # Clear the spinner line
        print(f"\r{' ' * (len(message) + 35)}\r", end="", flush=True)

    thread = threading.Thread(target=_spin, daemon=True)
    thread.start()
    try:
        yield
    finally:
        stop_event.set()
        thread.join()


def build_http_session(headers: dict[str, str]):
    import requests

    session = requests.Session()
    session.headers.update(headers)
    return session


def normalize_connection_kwargs(connection_kwargs: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(connection_kwargs)
    normalized["http_scheme"] = "https"
    normalized.setdefault("port", 8080)

    host = normalized.get("host")
    if not host:
        return normalized

    raw_host = str(host).strip()
    parsed = urlparse(raw_host if "://" in raw_host else f"https://{raw_host}")
    if parsed.hostname:
        normalized["host"] = parsed.hostname
    if parsed.port is not None:
        normalized["port"] = parsed.port

    return normalized


def build_engine_url(host: str, port: int, user: str, catalog: str, schema: str) -> str:
    return f"trino://{user}@{host}:{port}/{catalog}/{schema}"


def build_full_table_name(table: str) -> str:
    return f"{SRC_CATALOG}.{SRC_SCHEMA}.{table}"


def build_describe_sql(table: str) -> str:
    return f"DESCRIBE {build_full_table_name(table)}"


def build_bcv_table_name(table: str) -> str:
    return f"{BCV_CATALOG}.{BCV_SCHEMA}.{table}"


def build_bcv_describe_sql(table: str) -> str:
    return f"DESCRIBE {build_bcv_table_name(table)}"


def sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def build_usage_sql(table: str, column: str) -> str:
    return build_usage_sql_batch(table, [column])


def build_usage_sql_batch(table: str, columns: list[str]) -> str:
    col_list = ", ".join(sql_literal(col) for col in columns)
    return f"""
SELECT
    a.col AS column_name,
    CASE
        WHEN a.user IN ('sa-dataapp-insights', 'sa-dmo-aqs', 'sa-dataapp-yield', 'svc-ciec-sct', 'sa-trust_standards', 'sa-analytics-scrum') THEN 'Insights'
        WHEN a.source LIKE '%Arena%' THEN 'Arena'
        WHEN a.source LIKE '%lqs%' THEN 'LQS'
        ELSE 'Others'
    END AS user,
    count(*) as usage_count
FROM (
    SELECT DISTINCT
        q.pid,
        q.user,
        q.source,
        q.create_date,
        pcu.catalog,
        pcu.schema,
        pcu.table_name,
        p.col
    FROM etl.pgw_etl."query" q
    JOIN etl.pgw_etl."presto_column_usage" pcu on pcu.pid = q.pid
    CROSS JOIN UNNEST(pcu.columns) AS p(col)
    WHERE q.create_date >= DATE '2026-01-01'
    AND pcu.create_date >= DATE '2026-01-01'
    AND pcu.catalog = {sql_literal(SRC_CATALOG)}
    AND pcu.schema = {sql_literal(SRC_SCHEMA)}
    AND pcu.table_name = {sql_literal(table)}
    AND env = 'prd'
    AND q.user NOT IN ('sqyang', 'yjgou', 'kbhargava', 'yuwang', 'zhfan')
    AND q.source NOT IN ('presto-python-client')
    AND p.col IN ({col_list})
    AND (NOT regexp_like(q.query, '(?i)select\\s*\\*'))
    AND error_type is null AND error_name is null
) a
GROUP BY 1, 2
ORDER BY 1, 3 DESC
""".strip()


def build_connect_args(connection_kwargs: dict[str, Any]) -> dict[str, Any]:
    headers = connection_kwargs.get("http_headers") or {}
    return {
        "http_scheme": connection_kwargs.get("http_scheme", "https"),
        "http_session": build_http_session(headers),
        "request_timeout": connection_kwargs.get("request_timeout", 5.0),
        "max_attempts": 1,
    }


def format_rows(rows: list[tuple[Any, ...]], description: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    columns = [column[0] for column in description]
    return [dict(zip(columns, row)) for row in rows]


def remove_describe_metadata_columns(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {name: value for name, value in row.items() if name.lower() not in EXCLUDED_DESCRIBE_COLUMNS}
        for row in rows
    ]


def get_case_insensitive(row: dict[str, Any], key: str) -> Any:
    for name, value in row.items():
        if name.lower() == key.lower():
            return value
    return ""


def normalize_field_name_for_size_lookup(field_name: str) -> str:
    return field_name.replace("__", ".")


def lookup_size(src_field: str, field_sizes: dict[str, float]) -> float | str:
    if not src_field:
        return ""

    size = field_sizes.get(normalize_field_name_for_size_lookup(src_field))
    if size is None:
        return ""
    return round(size, 2)


def load_field_sizes(table: str, field_size_dir: Path = FIELD_SIZE_DIR) -> dict[str, float]:
    from openpyxl import load_workbook

    workbook_path = field_size_dir / f"{table}_raw_size_in_TiB.xlsx"
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(rows)]
    field_index = headers.index("Field Name")
    size_index = headers.index("Size (TiB)")

    field_sizes = {}
    for row in rows:
        field_name = row[field_index]
        size = row[size_index]
        if field_name and size is not None:
            field_sizes[str(field_name)] = float(size)
    return field_sizes


def load_etl_fields(table: str, etl_fields_path: Path = ETL_FIELDS_PATH) -> set[str]:
    if not etl_fields_path.exists():
        return set()

    data = json.loads(etl_fields_path.read_text(encoding="utf-8"))
    return {str(field) for field in data.get(table, [])}


def add_etl_usage_info(
    table: str,
    rows: list[dict[str, Any]],
    etl_fields_path: Path = ETL_FIELDS_PATH,
) -> list[dict[str, Any]]:
    etl_fields = load_etl_fields(table, etl_fields_path)
    for row in rows:
        row.setdefault("usage:ETL", "")
        if not is_missing_bcv_column(row):
            continue
        normalized_src_field = normalize_field_name_for_size_lookup(str(row.get("src_field") or ""))
        if normalized_src_field in etl_fields:
            row["usage:ETL"] = "Y"
    return rows


def compare_schema_rows(
    table: str,
    src_rows: list[dict[str, Any]],
    bcv_rows: list[dict[str, Any]],
    field_sizes: dict[str, float] | None = None,
) -> list[dict[str, Any]]:
    field_sizes = field_sizes or {}
    bcv_by_name = {get_case_insensitive(row, "column"): row for row in bcv_rows}
    compared = []

    for src_row in src_rows:
        src_name = get_case_insensitive(src_row, "column")
        src_type = get_case_insensitive(src_row, "type")
        bcv_row = bcv_by_name.get(src_name)
        bcv_type = get_case_insensitive(bcv_row, "type") if bcv_row else ""
        compared.append(
            {
                "status": "MATCHED" if bcv_row and src_type == bcv_type else "DIFF",
                "src_field": src_name,
                "src_type": src_type,
                "bcv_field": get_case_insensitive(bcv_row, "column") if bcv_row else "",
                "bcv_type": bcv_type,
                "size": lookup_size(src_name, field_sizes),
            }
        )

    src_names = {get_case_insensitive(row, "column") for row in src_rows}
    for bcv_row in bcv_rows:
        bcv_name = get_case_insensitive(bcv_row, "column")
        if bcv_name in src_names:
            continue
        compared.append(
            {
                "status": "DIFF",
                "src_field": "",
                "src_type": "",
                "bcv_field": bcv_name,
                "bcv_type": get_case_insensitive(bcv_row, "type"),
                "size": "",
            }
        )

    return compared


def count_missing_bcv_columns(rows: list[dict[str, Any]]) -> int:
    return sum(
        1
        for row in rows
        if row.get("status") == "DIFF" and not row.get("bcv_field") and not row.get("bcv_type")
    )


def is_missing_bcv_column(row: dict[str, Any]) -> bool:
    return (
        row.get("status") == "DIFF"
        and bool(row.get("src_field"))
        and not row.get("bcv_field")
        and not row.get("bcv_type")
    )


def add_usage_info(
    table: str,
    rows: list[dict[str, Any]],
    connection_kwargs: dict[str, Any],
    limit: int = USAGE_QUERY_LIMIT,
) -> list[dict[str, Any]]:
    for row in rows:
        for column in USAGE_COLUMNS:
            row.setdefault(column, "")

    missing_rows = [row for row in rows if is_missing_bcv_column(row)][:limit]
    total = len(missing_rows)
    if total == 0:
        return rows

    row_by_field: dict[str, dict[str, Any]] = {str(r["src_field"]): r for r in missing_rows}
    num_batches = (total + USAGE_QUERY_BATCH_SIZE - 1) // USAGE_QUERY_BATCH_SIZE

    for batch_idx in range(num_batches):
        batch_start = batch_idx * USAGE_QUERY_BATCH_SIZE
        batch_end = min(batch_start + USAGE_QUERY_BATCH_SIZE, total)
        batch_columns = [str(r["src_field"]) for r in missing_rows[batch_start:batch_end]]

        usage_rows: list[dict[str, Any]] = []
        for attempt in range(USAGE_QUERY_MAX_RETRIES + 1):
            try:
                with spinning_cursor(
                    f"Retrieving usage info for missing columns in BCV:"
                    f" batch {batch_idx + 1}/{num_batches} (columns {batch_start + 1}–{batch_end} of {total})"
                ):
                    usage_rows = execute_sql(
                        build_usage_sql_batch(table, batch_columns),
                        connection_kwargs=connection_kwargs,
                    )
                break
            except Exception as exc:
                if attempt < USAGE_QUERY_MAX_RETRIES:
                    log_info(
                        f"Retry {attempt + 1}/{USAGE_QUERY_MAX_RETRIES} for batch"
                        f" {batch_idx + 1}/{num_batches}: {exc}"
                    )
                else:
                    log_info(
                        f"Failed to retrieve usage info for batch {batch_idx + 1}/{num_batches}"
                        f" after {USAGE_QUERY_MAX_RETRIES} retries: {exc}"
                    )

        for usage_row in usage_rows:
            col_name = str(get_case_insensitive(usage_row, "column_name"))
            user = get_case_insensitive(usage_row, "user")
            usage_column = f"usage:{user}"
            if col_name in row_by_field and usage_column in USAGE_COLUMNS:
                row_by_field[col_name][usage_column] = get_case_insensitive(usage_row, "usage_count")

    return rows, missing_rows


def execute_sql(
    sql: str,
    connection_kwargs: dict[str, Any],
    engine_factory: Callable[..., Any] | None = None,
) -> list[dict[str, Any]]:
    kwargs = normalize_connection_kwargs(connection_kwargs)
    if engine_factory is None:
        from sqlalchemy import create_engine

        engine_factory = create_engine

    engine = engine_factory(
        build_engine_url(
            host=kwargs["host"],
            port=kwargs["port"],
            user=kwargs["user"],
            catalog=kwargs["catalog"],
            schema=kwargs["schema"],
        ),
        connect_args=build_connect_args(kwargs),
    )

    connection = engine.raw_connection()
    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute(sql)
        rows = cursor.fetchall()
        return format_rows([tuple(row) for row in rows], cursor.description)
    finally:
        if cursor is not None:
            try:
                cursor.close()
            except Exception as exc:
                print(f"Warning: failed to close cursor cleanly: {exc}", file=sys.stderr)
        try:
            connection.close()
        except Exception as exc:
            print(f"Warning: failed to close connection cleanly: {exc}", file=sys.stderr)


def print_rows_as_json(rows: list[dict[str, Any]]) -> None:
    print(json.dumps(rows, ensure_ascii=False, indent=2, default=str))


def write_rows_as_json_file(rows: list[dict[str, Any]], table: str) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / f"{table}.json").write_text(
        json.dumps(rows, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )


def write_rows_as_csv_file(rows: list[dict[str, Any]], filename: str, fieldnames: tuple[str, ...]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with (OUTPUT_DIR / filename).open("w", encoding="utf-8", newline="") as output:
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_connection_kwargs(
    host: str,
    port: int,
    user: str,
    catalog: str,
    schema: str,
    request_timeout: float,
    auth_token: str | None,
    auth_header: str | None,
) -> dict[str, Any]:
    return {
        "host": host,
        "port": port,
        "user": user,
        "catalog": catalog,
        "schema": schema,
        "request_timeout": request_timeout,
        "http_headers": build_http_headers(auth_token, auth_header),
    }


def retrieve_column_list(
    table_name: str,
    describe_sql: str,
    connection_kwargs: dict[str, Any],
) -> list[dict[str, Any]]:
    log_info(f"Retrieving column list for table: {table_name}")
    with spinning_cursor():
        result = execute_sql(describe_sql, connection_kwargs=connection_kwargs)
    return result


def resolve_table(table: str | None) -> str:
    if table:
        if table not in ALLOWED_TABLES:
            raise SystemExit(f"Invalid table: {table}")
        return table

    UI_CONSOLE.print(
        Panel(
            f"{APP_BANNER}\nBCV Analyzer {APP_VERSION}\nUse ↑/↓ to choose a table, then press Enter.\n\n"
            "[bold]Analysis Summary Rules[/bold] (applied to DIFF rows where SRC has value, BCV is missing):\n"
            "  [bold cyan]■ Recommended for Backfill   [/bold cyan]: usage (ETL = [bold]Y[/bold] OR Insights [bold]> 0[/bold] OR Arena [bold]> 0[/bold] OR LQS [bold]≥ 10[/bold] OR Others [bold]≥ 100[/bold])  AND  size < 0.03 TiB (or unknown)\n"
            "  [bold yellow]■ Recommended Excluded       [/bold yellow]: usage (ETL = [bold]Y[/bold] OR Insights [bold]> 0[/bold] OR Arena [bold]> 0[/bold] OR LQS [bold]≥ 10[/bold] OR Others [bold]≥ 100[/bold])  AND  size ≥ 0.03 TiB\n"
            "  [bold red]■ Recommended No Backfill    [/bold red]: usage below threshold  (ETL is blank, Insights = 0, Arena = 0, LQS < 10, Others < 100)",
            title="BCV Analyzer",
            border_style="cyan",
        )
    )
    selected = questionary.select("Select table:", choices=list(ALLOWED_TABLES)).ask()
    if not selected:
        raise SystemExit("No table selected")
    return selected


def resolve_run_mode() -> RunMode:
    selected = questionary.select(
        "Select run mode:",
        choices=[
            questionary.Choice(
                title=f"Trial  (query Presto for first {USAGE_QUERY_LIMIT} missing columns only)",
                value=RunMode.TRIAL,
            ),
            questionary.Choice(
                title="Full Run  (query Presto for ALL missing columns, takes longer)",
                value=RunMode.FULL_RUN,
            ),
        ],
    ).ask()
    if not selected:
        raise SystemExit("No run mode selected")
    return selected


def validate_connection_args(host: str | None, user: str | None) -> None:
    required = {
        "host": host,
        "user": user,
    }
    missing = [name for name, value in required.items() if not value]
    if missing:
        raise SystemExit(f"Missing required connection args: {', '.join(missing)}")


def get_usage_int(row: dict[str, Any], column: str) -> int:
    try:
        return int(row.get(column) or 0)
    except (ValueError, TypeError):
        return 0


def is_etl_used(row: dict[str, Any]) -> bool:
    return row.get("usage:ETL") == "Y"


def usage_meets_threshold(row: dict[str, Any]) -> bool:
    return (
        is_etl_used(row)
        or get_usage_int(row, "usage:Insights") > 0
        or get_usage_int(row, "usage:Arena") > 0
        or get_usage_int(row, "usage:LQS") >= 10
        or get_usage_int(row, "usage:Others") >= 100
    )


def get_recommended_action(row: dict[str, Any], was_queried: bool) -> str:
    if not is_missing_bcv_column(row):
        return ""
    if not was_queried and not is_etl_used(row):
        return ""
    if usage_meets_threshold(row):
        size = row.get("size")
        if not isinstance(size, (int, float)) or float(size) < 0.03:
            return "Backfill"
        return "Excluded - Size Too Large"
    return "No Backfill - Low Usage"


def print_analysis_summary(queried_rows: list[dict[str, Any]]) -> None:
    from rich.table import Table

    # All queried_rows already satisfy is_missing_bcv_column
    recommended = [
        row for row in queried_rows
        if usage_meets_threshold(row)
        and (
            not isinstance(row.get("size"), (int, float))
            or float(row["size"]) < 0.03
        )
    ]

    excluded = [
        row for row in queried_rows
        if usage_meets_threshold(row)
        and isinstance(row.get("size"), (int, float))
        and float(row["size"]) >= 0.03
    ]

    low_usage = [
        row for row in queried_rows
        if not usage_meets_threshold(row)
    ]

    UI_CONSOLE.print()

    # --- Blue: recommended for backfill ---
    if recommended:
        rec_table = Table(show_header=True, header_style="bold cyan", border_style="cyan", show_lines=True)
        rec_table.add_column("#", justify="right", style="dim", no_wrap=True)
        rec_table.add_column("Column Name", style="bold white", no_wrap=True)
        rec_table.add_column("Size (TiB)", justify="right", style="green", no_wrap=True)
        for column in USAGE_COLUMNS:
            rec_table.add_column(column, justify="right", no_wrap=True)
        for i, row in enumerate(recommended, 1):
            rec_table.add_row(
                str(i), str(row["src_field"]), str(row["size"]),
                *(str(row.get(column) or "") for column in USAGE_COLUMNS),
            )
        UI_CONSOLE.print(Panel(
            rec_table,
            title=f"[bold cyan]Recommended for Backfill — {len(recommended)} column(s)[/bold cyan]",
            border_style="cyan",
        ))
    else:
        UI_CONSOLE.print(Panel(
            "No columns recommended for backfill.",
            title="[bold cyan]Recommended for Backfill[/bold cyan]",
            border_style="cyan",
        ))

    # --- Yellow: usage meets threshold but size too large ---
    if excluded:
        exc_table = Table(show_header=True, header_style="bold yellow", border_style="yellow", show_lines=True)
        exc_table.add_column("#", justify="right", style="dim", no_wrap=True)
        exc_table.add_column("Column Name", style="white", no_wrap=True)
        exc_table.add_column("Size (TiB)", justify="right", style="yellow", no_wrap=True)
        for column in USAGE_COLUMNS:
            exc_table.add_column(column, justify="right", no_wrap=True)
        for i, row in enumerate(excluded, 1):
            exc_table.add_row(
                str(i), str(row["src_field"]), str(row["size"]),
                *(str(row.get(column) or "") for column in USAGE_COLUMNS),
            )
        UI_CONSOLE.print(Panel(
            exc_table,
            title=f"[bold yellow]Recommended Excluded (size ≥ 0.03 TiB) — {len(excluded)} column(s)[/bold yellow]",
            border_style="yellow",
        ))

    # --- Red: usage below threshold, no backfill ---
    if low_usage:
        low_table = Table(show_header=True, header_style="bold red", border_style="red", show_lines=True)
        low_table.add_column("#", justify="right", style="dim", no_wrap=True)
        low_table.add_column("Column Name", style="white", no_wrap=True)
        low_table.add_column("Size (TiB)", justify="right", no_wrap=True)
        for column in USAGE_COLUMNS:
            low_table.add_column(column, justify="right", no_wrap=True)
        for i, row in enumerate(low_usage, 1):
            low_table.add_row(
                str(i), str(row["src_field"]), str(row["size"]),
                *(str(row.get(column) or "") for column in USAGE_COLUMNS),
            )
        UI_CONSOLE.print(Panel(
            low_table,
            title=f"[bold red]Recommended No Backfill (usage below threshold) — {len(low_usage)} column(s)[/bold red]",
            border_style="red",
        ))


def main(
    host: Annotated[str | None, typer.Option(help="Presto host")] = os.getenv("PRESTO_HOST"),
    port: Annotated[int, typer.Option(help="Presto port")] = int(os.getenv("PRESTO_PORT", "8080")),
    user: Annotated[str | None, typer.Option(help="Presto user")] = os.getenv("PRESTO_USER"),
    request_timeout: Annotated[
        float,
        typer.Option("--request-timeout", help="Request timeout in seconds"),
    ] = float(os.getenv("PRESTO_REQUEST_TIMEOUT", "5")),
    auth_token: Annotated[
        str | None,
        typer.Option("--auth-token", help="Presto gateway token"),
    ] = os.getenv("PRESTO_AUTH_TOKEN"),
    auth_header: Annotated[
        str | None,
        typer.Option("--auth-header", help="Auth header name"),
    ] = os.getenv("PRESTO_AUTH_HEADER"),
    table: Annotated[str | None, typer.Option("--table", help="BCV table name")] = None,
) -> None:
    selected_table = resolve_table(table)
    run_mode = resolve_run_mode()
    log_info(f"Run mode: {run_mode.value}")
    validate_connection_args(host, user)
    src_table_name = build_full_table_name(selected_table)
    bcv_table_name = build_bcv_table_name(selected_table)
    src_connection_kwargs = build_connection_kwargs(
        host,
        port,
        user,
        SRC_CATALOG,
        SRC_SCHEMA,
        request_timeout,
        auth_token,
        auth_header,
    )
    bcv_connection_kwargs = build_connection_kwargs(
        host,
        port,
        user,
        BCV_CATALOG,
        BCV_SCHEMA,
        request_timeout,
        auth_token,
        auth_header,
    )
    rows = retrieve_column_list(
        src_table_name,
        f"DESCRIBE {src_table_name}",
        src_connection_kwargs,
    )
    bcv_rows = retrieve_column_list(
        bcv_table_name,
        f"DESCRIBE {bcv_table_name}",
        bcv_connection_kwargs,
    )
    src_output_rows = remove_describe_metadata_columns(rows)
    bcv_output_rows = remove_describe_metadata_columns(bcv_rows)
    log_info("Retrieving column size")
    field_sizes = load_field_sizes(selected_table, FIELD_SIZE_DIR)
    write_rows_as_json_file(src_output_rows, selected_table)
    write_rows_as_json_file(bcv_output_rows, f"bcv_{selected_table}")
    comparison_rows = compare_schema_rows(selected_table, src_output_rows, bcv_output_rows, field_sizes)
    add_etl_usage_info(selected_table, comparison_rows, ETL_FIELDS_PATH)
    usage_limit = USAGE_QUERY_LIMIT if run_mode == RunMode.TRIAL else sys.maxsize
    comparison_rows, queried_rows = add_usage_info(selected_table, comparison_rows, src_connection_kwargs, limit=usage_limit)
    queried_fields = {str(r["src_field"]) for r in queried_rows}
    summary_rows = list(queried_rows)
    summary_fields = set(queried_fields)
    for row in comparison_rows:
        src_field = str(row.get("src_field") or "")
        if is_missing_bcv_column(row) and is_etl_used(row) and src_field not in summary_fields:
            summary_rows.append(row)
            summary_fields.add(src_field)
    for row in comparison_rows:
        was_queried = is_missing_bcv_column(row) and str(row.get("src_field")) in queried_fields
        row["recommended_action"] = get_recommended_action(row, was_queried)
    print_analysis_summary(summary_rows)
    result_filename = f"{selected_table}_result.csv"
    write_rows_as_csv_file(
        comparison_rows,
        result_filename,
        ("status", "src_field", "src_type", "bcv_field", "bcv_type", "size", *USAGE_COLUMNS, "recommended_action"),
    )
    result_path = (OUTPUT_DIR / result_filename).resolve()
    UI_CONSOLE.print(f"[{current_timestamp()}] Completed! results are written to [bold green]{result_path}[/bold green]")


if __name__ == "__main__":
    typer.run(main)
